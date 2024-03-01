import json
import openpyxl
wb = openpyxl.load_workbook('GBF素材需求量 (1).xlsx')     # 開啟 Excel 檔案

s1 = wb['工作表6']        # 取得工作表名稱為「工作表1」的內容
s2 = wb.active           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

print(s1.title, s1.max_row, s1.max_column)  # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
print(s2.title, s2.max_row, s2.max_column)  # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數

print(s1.sheet_properties)   # 印出工作表屬性

def search(json_str, no):
    return [datum['quantity'] for datum in json.loads(json_str) if datum['id']==no]

jsonFile = open('20240301140423_yan.json','r')
f =  jsonFile.read()   # 要先使用 read 讀取檔案
a = json.loads(f)      # 再使用 loads
print (a[0]['id'])
with open('ids.json','r') as f:
    data = json.load(f)
    for i in range(data.__len__()):
        print(search(json.dumps(a), data[i]))
        #print(data[i])
        s1['B'+(str(i+1))].value=search(json.dumps(a), data[i])[0]
wb.save('GBF素材需求量 (1).xlsx')  # 存檔